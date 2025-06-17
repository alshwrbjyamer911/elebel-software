import os
import openpyxl

EXCEL_FILE = "lora_nodes.xlsx"
BAND = "868000000"
NETWORKID_START = 1
NETWORKID_END = 16
CPIN_PASSWORD = "LoRaSecure1234"

MAX_ADDRESS = 65534  # Avoid 0, reserve 65535 if needed

def load_existing_data():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["NetworkID", "Device Address", "Password", "Band"])
        wb.save(EXCEL_FILE)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    return wb, wb.active

def get_used_addresses(ws, netid):
    return {row[1] for row in ws.iter_rows(min_row=2, values_only=True) if row[0] == netid}

def main():
    wb, ws = load_existing_data()

    for netid in range(NETWORKID_START, NETWORKID_END + 1):
        print(f"Processing NETWORKID={netid}...")
        used = get_used_addresses(ws, netid)
        available = [addr for addr in range(1, MAX_ADDRESS + 1) if addr not in used]

        if not available:
            print(f"All addresses used for NETWORKID={netid}, moving on.")
            continue

        for addr in available:
            ws.append([netid, addr, CPIN_PASSWORD, BAND])
            print(f"Assigned: NETWORKID={netid}, ADDRESS={addr}")

        # Optional: Save after each NETWORKID
        wb.save(EXCEL_FILE)

    print("✅ Generation complete. Data saved to", EXCEL_FILE)
    wb.save(EXCEL_FILE)

if __name__ == "__main__":
    main()
